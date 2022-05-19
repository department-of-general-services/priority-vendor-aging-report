from datetime import datetime

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from dgs_fiscal.etl.prompt_payment.constants import DIVISIONS, VALIDATION


def compute_age_of_invoice(df: pd.DataFrame) -> pd.Series:
    """Calculates 'Age of Invoice' field"""
    now = datetime.now().date()
    age_of_invoice = now - df["Invoice Date"].dt.date
    return age_of_invoice


def compute_days_outstanding(df: pd.DataFrame) -> pd.Series:
    """Calculates 'Days Outstanding' field"""
    bins = [0, 30, 60, 90, 120, 10000]
    labels = [
        "Less than 30",
        "30 to 60 days",
        "60 to 90 days",
        "90 to 120 days",
        "Over 120 days",
    ]
    col = df["Days Since Creation"]
    days_outstanding = pd.cut(col, bins, False, labels)
    return days_outstanding


def compute_days_with_baps(df: pd.DataFrame) -> pd.DataFrame:
    """Calculates 'Days with BAPS' field"""
    bins = [0, 10, 10000]
    labels = [
        "less than 10 days",
        "at least 10 days",
    ]
    col = df["Days Since Creation"] - df["Status Age (days)"]
    days_with_baps = pd.cut(col, bins, False, labels)
    return days_with_baps


def update_division(  # pylint: disable=dangerous-default-value
    df: pd.DataFrame,
    staff_mapping: list = DIVISIONS,
) -> pd.DataFrame:
    """Assigns the division associated with a give AP staff"""

    # change default to "No matching division"
    df["Division"] = "No matching division"

    # update rows that have multiple people listed
    multiples = df["DGS Name"].str.contains(",")
    df.loc[multiples, "Division"] = "Multiple people listed"

    # update remaining rows based on staff mapping
    for division, names in staff_mapping.items():
        staff = df["DGS Name"].isin(names)
        df.loc[staff, "Division"] = division

    return df


def format_workbook(workbook, sheet="Prompt Payment"):
    """Formats workbook before saving it. Called by overwrite_sheet() function
    Args:
        workbook (openpyxl.workbook): Workbook object of OneDrive excel file
        sheet (str): Name of the sheet with the Prompt Payment report
        to_loc (int): Position of the Prompt Payment sheet after formatting
    Returns:
        message (str): Message indicating a successful formatting
        workbook (openpyxl.workbook) Formatted workbook object
    """
    # load Prompt Payment worksheet
    worksheet = workbook[sheet]

    # add conditional highlighting
    add_conditional_formatting(worksheet, "Age of Invoice", 15, 30)
    add_conditional_formatting(worksheet, "Days Since Creation", 15, 30)
    add_conditional_formatting(worksheet, "Status Age (days)", 7, 15)

    # hide Days Since Creation
    col = get_excel_range(worksheet, "Days Since Creation")[0]
    worksheet.column_dimensions[col].hidden = True

    # add dropdown to Current Status column
    add_data_validation(worksheet, "Current Status")

    # adjust column width, filter worksheet, and freeze header row
    autoresize_cols(worksheet)
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"


def add_conditional_formatting(ws, column, low, high):
    """Adds conditional formatting to a column in a worksheet
    Args:
        ws (worksheet): Worksheet object for the Prompt Payment report
        column (str): Name of the column to apply the conditional formatting to
        low (int): The cutoff between green and yellow highlighting
        high (int): The cutoff between yellow and red highlighting
    """

    range = get_excel_range(ws, column)

    # helper function to set the fill format
    def fill(name):
        colors = {"green": "B6D7A8", "yellow": "FFF2CC", "red": "E6B8AF"}
        format = PatternFill(
            start_color=colors[name], end_color=colors[name], fill_type="solid"
        )
        return format

    # helper function to create the conditional formatting rule
    def rule(operator, value, color):
        condition = CellIsRule(
            operator=(operator),
            formula=[str(value)],
            stopIfTrue=True,
            fill=fill(color),
        )
        return condition

    ws.conditional_formatting.add(range, rule("lessThan", low, "green"))
    ws.conditional_formatting.add(range, rule("lessThan", high, "yellow"))
    ws.conditional_formatting.add(
        range, rule("greaterThanOrEqual", high, "red")
    )


def add_data_validation(ws, column, src=VALIDATION):
    """Adds data validation to a column in a worksheet
    Args:
        ws (worksheet): Worksheet object for the Prompt Payment report
        column (str): Name of the column to apply the data validation to
        src_sheet (str): Name of the sheet that contains the validation options
        src_range (str): Cell range for options on src_sheet
    """

    range = get_excel_range(ws, column)
    options = ",".join(src)
    options_str = f'"{options}"'

    # create the validation rule
    validation = DataValidation(
        type="list",
        formula1=options_str,
        allow_blank=True,
    )
    validation.add(range)

    # add the validation to the seet
    ws.add_data_validation(validation)


def autoresize_cols(ws, max_width=80, min_width=10):
    """Adjusts column width to fit column contents upt to a max value
    Args:
        ws (worksheet): Worksheet object for the Prompt Payment report
        max_width (int): The maximum width for a column
    """

    def cell_len(cell):
        return len(str(cell.value))

    for col in ws.columns:
        column = col[0].column_letter  # Get the column name
        max_cell = max(col, key=cell_len)
        adjusted_width = cell_len(max_cell) + 2
        # ensure width remains within limits
        width = sorted([min_width, adjusted_width, max_width])[1]
        ws.column_dimensions[column].width = width


def get_excel_range(ws, name):
    """Get the excel range for a column"""
    headers = [cell.value for cell in ws[1]]
    col = get_column_letter(headers.index(name) + 1)
    return f"{col}2:{col}1048576"
